from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sqlite3
import tempfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

import openpyxl
import polars as pl


SOURCE_URL = (
    "https://cdn1.gopublic.dk/sundhedsdatastyrelsen/Media/638999138200017589/"
    "Register%20for%20Udvalgte%20Kroniske%20Sygdomme%20og%20Sv%C3%A6re%20Psykiske%20"
    "Lidelser%20(RUKS)%202010-2025%20(udgivet%2028.%20november%202025).XLSX"
)
YEAR_COLUMNS = list(range(2010, 2026))
AGE_GROUP_ORDER = [
    "Standardiseret",
    "0-4",
    "5-9",
    "10-14",
    "15-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65-69",
    "70-74",
    "75-79",
    "80-84",
    "85+",
    "Alle aldre",
]
AGE_GROUP_ORDER_MAP = {label: index for index, label in enumerate(AGE_GROUP_ORDER, start=1)}


@dataclass
class SourceMetadata:
    workbook_title: str
    source_release_date_text: str
    source_release_date_iso: str
    department: str
    source_url: str
    source_sha256: str
    downloaded_at_utc: str


def slugify(value: str) -> str:
    replacements = {
        "æ": "ae",
        "ø": "oe",
        "å": "aa",
        "Æ": "ae",
        "Ø": "oe",
        "Å": "aa",
    }
    result = value.strip()
    for source, target in replacements.items():
        result = result.replace(source, target)
    pieces = []
    for character in result.lower():
        if character.isalnum():
            pieces.append(character)
        else:
            pieces.append("-")
    slug = "".join(pieces)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-")


def parse_danish_date(value: str) -> str:
    months = {
        "januar": 1,
        "februar": 2,
        "marts": 3,
        "april": 4,
        "maj": 5,
        "juni": 6,
        "juli": 7,
        "august": 8,
        "september": 9,
        "oktober": 10,
        "november": 11,
        "december": 12,
    }
    cleaned = value.strip().replace(".", "")
    day_text, month_text, year_text = cleaned.split()
    return f"{int(year_text):04d}-{months[month_text.lower()]:02d}-{int(day_text):02d}"


def normalize_sex(value: str) -> tuple[str, str]:
    mapping = {
        "Begge": ("both", "Begge"),
        "Kvinder": ("women", "Kvinder"),
        "Mænd": ("men", "Mænd"),
        "Standardiseret": ("standardized", "Standardiseret"),
    }
    return mapping.get(value, (slugify(value), value))


def normalize_measure(value: str) -> tuple[str, str]:
    mapping = {
        "Incidens (nye sygdomstilfælde)": ("incidence", value),
        "Prævalens (sygdomsforekomst)": ("prevalence", value),
    }
    return mapping.get(value, (slugify(value), value))


def normalize_geo_level(value: str) -> str:
    mapping = {
        "Hele landet": "country",
        "Regioner": "region",
        "Kommuner": "municipality",
    }
    return mapping.get(value, slugify(value))


def normalize_unit(value: str) -> tuple[str, str, str]:
    mapping = {
        "Antal personer med sygdom": ("count", "persons", "none"),
        "Antal personer pr. 100.000 borgere": ("rate", "per_100k_population", "none"),
        "Antal personer aldersstandardiseret rate pr. 100.000": (
            "rate",
            "per_100k_population",
            "age_standardized",
        ),
        "Antal personer køns- og aldersstandardiseret rate pr. 100.000": (
            "rate",
            "per_100k_population",
            "sex_age_standardized",
        ),
    }
    return mapping.get(value, ("unknown", slugify(value), "unknown"))


def age_group_sort_key(value: str) -> int:
    return AGE_GROUP_ORDER_MAP.get(value, 999)


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def download_workbook(source_url: str, destination: Path) -> tuple[Path, str]:
    ensure_parent(destination)
    request = Request(source_url, headers={"User-Agent": "ruks-data-pipeline/0.1"})
    sha256 = hashlib.sha256()
    with urlopen(request) as response, destination.open("wb") as handle:
        while True:
            chunk = response.read(1024 * 1024)
            if not chunk:
                break
            sha256.update(chunk)
            handle.write(chunk)
    return destination, sha256.hexdigest()


def extract_cover_metadata(workbook: openpyxl.Workbook, source_url: str, source_sha256: str) -> SourceMetadata:
    sheet = workbook["FORSIDE"]
    workbook_title = str(sheet["C6"].value or "").strip()
    source_release_date_text = str(sheet["D7"].value or "").strip()
    department = str(sheet["D8"].value or "").strip()
    return SourceMetadata(
        workbook_title=workbook_title,
        source_release_date_text=source_release_date_text,
        source_release_date_iso=parse_danish_date(source_release_date_text),
        department=department,
        source_url=source_url,
        source_sha256=source_sha256,
        downloaded_at_utc=datetime.now(UTC).replace(microsecond=0).isoformat(),
    )


def create_sqlite_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA journal_mode = WAL;
        PRAGMA synchronous = NORMAL;

        CREATE TABLE IF NOT EXISTS dim_disease (
            disease_id INTEGER PRIMARY KEY,
            disease_slug TEXT NOT NULL UNIQUE,
            disease_label TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dim_geography (
            geography_id INTEGER PRIMARY KEY,
            geo_level TEXT NOT NULL,
            region_name TEXT,
            municipality_name TEXT,
            display_name TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dim_measure (
            measure_id INTEGER PRIMARY KEY,
            measure_code TEXT NOT NULL UNIQUE,
            measure_label TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dim_sex (
            sex_id INTEGER PRIMARY KEY,
            sex_code TEXT NOT NULL UNIQUE,
            sex_label TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dim_age_group (
            age_group_id INTEGER PRIMARY KEY,
            age_group_code TEXT NOT NULL UNIQUE,
            age_group_label TEXT NOT NULL,
            sort_order INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS dim_unit (
            unit_id INTEGER PRIMARY KEY,
            unit_code TEXT NOT NULL UNIQUE,
            source_label TEXT NOT NULL,
            value_kind TEXT NOT NULL,
            unit_label TEXT NOT NULL,
            standardization TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS fact_observation (
            observation_id INTEGER PRIMARY KEY,
            measure_id INTEGER NOT NULL,
            disease_id INTEGER NOT NULL,
            geography_id INTEGER NOT NULL,
            sex_id INTEGER NOT NULL,
            age_group_id INTEGER NOT NULL,
            unit_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            value REAL NOT NULL,
            source_release_date TEXT NOT NULL,
            source_url TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            FOREIGN KEY (measure_id) REFERENCES dim_measure (measure_id),
            FOREIGN KEY (disease_id) REFERENCES dim_disease (disease_id),
            FOREIGN KEY (geography_id) REFERENCES dim_geography (geography_id),
            FOREIGN KEY (sex_id) REFERENCES dim_sex (sex_id),
            FOREIGN KEY (age_group_id) REFERENCES dim_age_group (age_group_id),
            FOREIGN KEY (unit_id) REFERENCES dim_unit (unit_id)
        );
        """
    )


class DimensionCache:
    def __init__(self, connection: sqlite3.Connection, table_name: str, insert_sql: str) -> None:
        self.connection = connection
        self.table_name = table_name
        self.insert_sql = insert_sql
        self.cache: dict[tuple[Any, ...], int] = {}

    def get_or_create(self, key: tuple[Any, ...], values: tuple[Any, ...]) -> int:
        existing = self.cache.get(key)
        if existing is not None:
            return existing
        cursor = self.connection.execute(self.insert_sql, values)
        dimension_id = int(cursor.lastrowid)
        self.cache[key] = dimension_id
        return dimension_id


def write_history_row(history_path: Path, row: dict[str, Any]) -> None:
    ensure_parent(history_path)
    existing_hashes: set[str] = set()
    if history_path.exists():
        with history_path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for existing_row in reader:
                existing_hashes.add(existing_row["source_sha256"])
    if row["source_sha256"] in existing_hashes:
        return
    with history_path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "source_release_date",
                "release_tag",
                "source_sha256",
                "workbook_title",
                "source_row_count",
                "observation_count",
                "generated_at_utc",
            ],
        )
        writer.writerow(row)


def transform_hovedresultater(
    workbook: openpyxl.Workbook,
    metadata: SourceMetadata,
    repo_root: Path,
    artifacts_root: Path,
) -> dict[str, Any]:
    source_sheet = workbook["Hovedresultater"]
    output_csv_temp = artifacts_root / "build" / "ruks_hovedresultater_long.csv"
    output_csv_gz = artifacts_root / "releases" / "assets" / f"ruks_hovedresultater_long-{metadata.source_release_date_iso}.csv.gz"
    output_parquet = artifacts_root / "releases" / "assets" / f"ruks_hovedresultater_long-{metadata.source_release_date_iso}.parquet"
    output_sqlite = artifacts_root / "releases" / "assets" / f"ruks-{metadata.source_release_date_iso}.sqlite"
    release_tag_path = artifacts_root / "releases" / "release_tag.txt"
    release_title_path = artifacts_root / "releases" / "release_title.txt"
    release_notes_path = artifacts_root / "releases" / "release_notes.md"

    for path in [
        output_csv_temp,
        output_csv_gz,
        output_parquet,
        output_sqlite,
        release_tag_path,
        release_title_path,
        release_notes_path,
    ]:
        ensure_parent(path)

    # Recreate build outputs on each run so local reruns do not append to or
    # conflict with partially generated artifacts from an earlier execution.
    output_csv_temp.unlink(missing_ok=True)
    output_csv_gz.unlink(missing_ok=True)
    output_parquet.unlink(missing_ok=True)
    output_sqlite.unlink(missing_ok=True)
    output_sqlite.with_suffix(output_sqlite.suffix + "-shm").unlink(missing_ok=True)
    output_sqlite.with_suffix(output_sqlite.suffix + "-wal").unlink(missing_ok=True)

    connection = sqlite3.connect(output_sqlite)
    create_sqlite_schema(connection)

    disease_dim = DimensionCache(
        connection,
        "dim_disease",
        "INSERT INTO dim_disease (disease_slug, disease_label) VALUES (?, ?)",
    )
    geography_dim = DimensionCache(
        connection,
        "dim_geography",
        "INSERT INTO dim_geography (geo_level, region_name, municipality_name, display_name) VALUES (?, ?, ?, ?)",
    )
    measure_dim = DimensionCache(
        connection,
        "dim_measure",
        "INSERT INTO dim_measure (measure_code, measure_label) VALUES (?, ?)",
    )
    sex_dim = DimensionCache(
        connection,
        "dim_sex",
        "INSERT INTO dim_sex (sex_code, sex_label) VALUES (?, ?)",
    )
    age_group_dim = DimensionCache(
        connection,
        "dim_age_group",
        "INSERT INTO dim_age_group (age_group_code, age_group_label, sort_order) VALUES (?, ?, ?)",
    )
    unit_dim = DimensionCache(
        connection,
        "dim_unit",
        "INSERT INTO dim_unit (unit_code, source_label, value_kind, unit_label, standardization) VALUES (?, ?, ?, ?, ?)",
    )

    site_summary: dict[str, Any] = {
        "workbook_title": metadata.workbook_title,
        "source_release_date": metadata.source_release_date_text,
        "release_tag": f"ruks-{metadata.source_release_date_iso}",
        "diseases": set(),
        "series": {},
    }
    observation_count = 0
    source_row_count = 0
    fact_batch: list[tuple[Any, ...]] = []

    with output_csv_temp.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "measure_code",
                "measure_label",
                "disease_slug",
                "disease_label",
                "geo_level",
                "region_name",
                "municipality_name",
                "sex_code",
                "sex_label",
                "age_group_code",
                "age_group_label",
                "value_kind",
                "unit",
                "standardization",
                "source_unit_label",
                "year",
                "value",
                "source_release_date",
                "source_url",
                "source_sheet",
            ],
        )
        writer.writeheader()

        for row in source_sheet.iter_rows(min_row=4, values_only=True):
            if not any(value is not None for value in row):
                continue

            source_row_count += 1
            source_measure = str(row[0]).strip()
            disease_label = str(row[1]).strip()
            source_geo_level = str(row[2]).strip()
            region_name = str(row[3]).strip() if row[3] else None
            municipality_name = str(row[4]).strip() if row[4] else None
            source_unit_label = str(row[5]).strip()
            source_sex = str(row[6]).strip()
            source_age_group = str(row[7]).strip()

            measure_code, measure_label = normalize_measure(source_measure)
            disease_slug = slugify(disease_label)
            geo_level = normalize_geo_level(source_geo_level)
            sex_code, sex_label = normalize_sex(source_sex)
            age_group_code = slugify(source_age_group)
            value_kind, unit, standardization = normalize_unit(source_unit_label)

            disease_id = disease_dim.get_or_create(
                (disease_slug,),
                (disease_slug, disease_label),
            )
            geography_display_name = (
                municipality_name
                or region_name
                or "Hele landet"
            )
            geography_id = geography_dim.get_or_create(
                (geo_level, region_name or "", municipality_name or ""),
                (geo_level, region_name, municipality_name, geography_display_name),
            )
            measure_id = measure_dim.get_or_create(
                (measure_code,),
                (measure_code, measure_label),
            )
            sex_id = sex_dim.get_or_create(
                (sex_code,),
                (sex_code, sex_label),
            )
            age_group_id = age_group_dim.get_or_create(
                (age_group_code,),
                (age_group_code, source_age_group, age_group_sort_key(source_age_group)),
            )
            unit_code = f"{value_kind}__{unit}__{standardization}"
            unit_id = unit_dim.get_or_create(
                (unit_code,),
                (unit_code, source_unit_label, value_kind, unit, standardization),
            )

            site_summary["diseases"].add(disease_label)

            for offset, year in enumerate(YEAR_COLUMNS, start=8):
                value = row[offset]
                if value is None:
                    continue

                observation_count += 1
                normalized_row = {
                    "measure_code": measure_code,
                    "measure_label": measure_label,
                    "disease_slug": disease_slug,
                    "disease_label": disease_label,
                    "geo_level": geo_level,
                    "region_name": region_name,
                    "municipality_name": municipality_name,
                    "sex_code": sex_code,
                    "sex_label": sex_label,
                    "age_group_code": age_group_code,
                    "age_group_label": source_age_group,
                    "value_kind": value_kind,
                    "unit": unit,
                    "standardization": standardization,
                    "source_unit_label": source_unit_label,
                    "year": year,
                    "value": value,
                    "source_release_date": metadata.source_release_date_iso,
                    "source_url": metadata.source_url,
                    "source_sheet": "Hovedresultater",
                }
                writer.writerow(normalized_row)

                fact_batch.append(
                    (
                        measure_id,
                        disease_id,
                        geography_id,
                        sex_id,
                        age_group_id,
                        unit_id,
                        year,
                        float(value),
                        metadata.source_release_date_iso,
                        metadata.source_url,
                        "Hovedresultater",
                    )
                )

                if (
                    geo_level == "country"
                    and sex_code == "both"
                    and age_group_code == "alle-aldre"
                    and standardization == "none"
                ):
                    summary_key = f"{disease_slug}::{measure_code}::{value_kind}"
                    series = site_summary["series"].setdefault(
                        summary_key,
                        {
                            "disease": disease_label,
                            "measure_code": measure_code,
                            "measure_label": measure_label,
                            "value_kind": value_kind,
                            "unit": unit,
                            "standardization": standardization,
                            "values": [],
                        },
                    )
                    series["values"].append({"year": year, "value": float(value)})

                if len(fact_batch) >= 10_000:
                    connection.executemany(
                        """
                        INSERT INTO fact_observation (
                            measure_id,
                            disease_id,
                            geography_id,
                            sex_id,
                            age_group_id,
                            unit_id,
                            year,
                            value,
                            source_release_date,
                            source_url,
                            source_sheet
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        fact_batch,
                    )
                    fact_batch.clear()

        if fact_batch:
            connection.executemany(
                """
                INSERT INTO fact_observation (
                    measure_id,
                    disease_id,
                    geography_id,
                    sex_id,
                    age_group_id,
                    unit_id,
                    year,
                    value,
                    source_release_date,
                    source_url,
                    source_sheet
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                fact_batch,
            )

    connection.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_fact_year ON fact_observation (year);
        CREATE INDEX IF NOT EXISTS idx_fact_measure ON fact_observation (measure_id);
        CREATE INDEX IF NOT EXISTS idx_fact_disease ON fact_observation (disease_id);
        CREATE INDEX IF NOT EXISTS idx_fact_geography ON fact_observation (geography_id);
        """
    )
    connection.commit()
    connection.close()

    lazy_frame = pl.scan_csv(output_csv_temp)
    lazy_frame.sink_parquet(output_parquet)

    with output_csv_temp.open("rb") as source_handle, tempfile.NamedTemporaryFile(delete=False) as temp_handle:
        with temp_handle:
            shutil.copyfileobj(source_handle, temp_handle)
        temp_path = Path(temp_handle.name)
    with temp_path.open("rb") as source_handle, output_csv_gz.open("wb") as target_handle:
        import gzip

        with gzip.GzipFile(fileobj=target_handle, mode="wb") as zipped_handle:
            shutil.copyfileobj(source_handle, zipped_handle)
    temp_path.unlink(missing_ok=True)
    output_csv_temp.unlink(missing_ok=True)

    manifest = {
        "workbook_title": metadata.workbook_title,
        "source_release_date_text": metadata.source_release_date_text,
        "source_release_date_iso": metadata.source_release_date_iso,
        "department": metadata.department,
        "source_url": metadata.source_url,
        "source_sha256": metadata.source_sha256,
        "downloaded_at_utc": metadata.downloaded_at_utc,
        "source_sheet": "Hovedresultater",
        "source_row_count": source_row_count,
        "observation_count": observation_count,
        "release_tag": f"ruks-{metadata.source_release_date_iso}",
        "notes": [
            "Blank cells in the source workbook are preserved as missing values.",
            "Interpretation of blank cells as disclosure suppression remains a documented follow-up task.",
        ],
        "artifacts": [
            output_csv_gz.name,
            output_parquet.name,
            output_sqlite.name,
        ],
    }

    manifest_dir = repo_root / "data" / "manifests"
    latest_manifest_path = manifest_dir / "latest.json"
    dated_manifest_path = manifest_dir / f"ruks-{metadata.source_release_date_iso}.json"
    ensure_parent(latest_manifest_path)
    latest_manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    dated_manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    site_summary["diseases"] = sorted(site_summary["diseases"])
    site_summary["source_row_count"] = source_row_count
    site_summary["observation_count"] = observation_count
    site_summary["series"] = list(site_summary["series"].values())
    site_summary_path = repo_root / "site" / "data" / "latest-summary.json"
    ensure_parent(site_summary_path)
    site_summary_path.write_text(json.dumps(site_summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    write_history_row(
        repo_root / "data" / "history" / "releases.csv",
        {
            "source_release_date": metadata.source_release_date_iso,
            "release_tag": f"ruks-{metadata.source_release_date_iso}",
            "source_sha256": metadata.source_sha256,
            "workbook_title": metadata.workbook_title,
            "source_row_count": source_row_count,
            "observation_count": observation_count,
            "generated_at_utc": metadata.downloaded_at_utc,
        },
    )

    release_tag_path.write_text(f"ruks-{metadata.source_release_date_iso}\n", encoding="utf-8")
    release_title_path.write_text(
        f"RUKS snapshot {metadata.source_release_date_iso}\n",
        encoding="utf-8",
    )
    release_notes_path.write_text(
        (
            f"# {metadata.workbook_title}\n\n"
            f"- Source release date: {metadata.source_release_date_text}\n"
            f"- Source SHA-256: `{metadata.source_sha256}`\n"
            f"- Rows in `Hovedresultater`: {source_row_count:,}\n"
            f"- Observed year-values: {observation_count:,}\n\n"
            "Assets in this release:\n\n"
            f"- `{output_csv_gz.name}`\n"
            f"- `{output_parquet.name}`\n"
            f"- `{output_sqlite.name}`\n"
        ),
        encoding="utf-8",
    )

    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Download and transform the public RUKS workbook.")
    parser.add_argument("--source-url", default=SOURCE_URL)
    parser.add_argument("--repo-root", default=Path.cwd(), type=Path)
    parser.add_argument("--artifacts-root", default=Path.cwd() / "artifacts", type=Path)
    parser.add_argument("--download-dir", default=Path.cwd() / "downloads", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = args.repo_root.resolve()
    artifacts_root = args.artifacts_root.resolve()
    download_dir = args.download_dir.resolve()
    download_path = download_dir / "ruks.xlsx"

    downloaded_file, source_sha256 = download_workbook(args.source_url, download_path)
    workbook = openpyxl.load_workbook(downloaded_file, read_only=True, data_only=True)
    metadata = extract_cover_metadata(workbook, args.source_url, source_sha256)
    transform_hovedresultater(workbook, metadata, repo_root=repo_root, artifacts_root=artifacts_root)
    workbook.close()
    return 0
